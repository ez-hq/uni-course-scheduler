#!/usr/bin/env python3
"""
University Course Planner — Excel Workbook Generator

Takes a combined JSON file (from LoomLoom cloud output or local Agent analysis)
and generates a 6-sheet .xlsx workbook following the ai-university-course-planner-schema.

Usage:
    python3 generate_excel.py <input.json> [--output <schedule.xlsx>]

The input JSON should have this structure:
{
    "meta": { "university": "...", "major": "...", ... },
    "course_overview": { "courses": [...] },
    "recommendations": { "recommendations": [...] },
    "weekly_schedule": { "schedule": [...], "daily_analysis": [...], ... }
}

If a section is missing, the corresponding sheet is created with headers only.
"""

import argparse
import json
import os
import sys
from datetime import datetime, date

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERROR: openpyxl is required. Install it with:")
    print("  pip install openpyxl")
    sys.exit(1)

# ─── Color Constants ───────────────────────────────────────────────────────

COLOR_DARK_RED = "8B0000"
COLOR_WHITE = "FFFFFF"
COLOR_GREEN = "228B22"
COLOR_BLUE = "2980B9"
COLOR_PURPLE = "8E44AD"
COLOR_LIGHT_BLUE = "85C1E9"
COLOR_ORANGE = "E67E22"
COLOR_RED = "FF0000"
COLOR_YELLOW = "F1C40F"
COLOR_GREY = "EAECEE"
COLOR_LIGHT_GREY = "D5D8DC"
COLOR_DARK_GREY = "7F8C8D"
COLOR_GOLD = "FFD700"
COLOR_TEAL = "117A65"
COLOR_CYAN = "1ABC9C"
COLOR_DARK_BLUE = "1A5276"
COLOR_LIGHT_RED = "FADBD8"
COLOR_GREEN_FILL = "27AE60"

NOT_FOUND = "— (not found, please add)"

# ─── Styling Helpers ───────────────────────────────────────────────────────


def make_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def make_font(color=COLOR_WHITE, bold=False, size=11, name="Calibri"):
    return Font(color=color, bold=bold, size=size, name=name)


def make_border(style="thin", color="000000"):
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_header_row(ws, row=1, cols=None):
    """Apply header styling to a row."""
    if cols is None:
        cols = ws.max_column
    header_fill = make_fill(COLOR_DARK_BLUE)
    header_font = make_font(color=COLOR_WHITE, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = make_border("thin", "808080")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[row].height = 30


def set_column_widths(ws, widths):
    """Set column widths from a list of (letter, width) tuples or a dict."""
    if isinstance(widths, dict):
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width
    elif isinstance(widths, list):
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width


def safe_val(val, default=NOT_FOUND):
    """Return val if not None/empty, else default. Converts lists/dicts to strings."""
    if val is None or val == "" or val == "NOT_FOUND":
        return default
    if isinstance(val, (list, dict)):
        if isinstance(val, list):
            joined = "; ".join(str(v) for v in val if v)
            return joined if joined else default
        return json.dumps(val, ensure_ascii=False)
    return val


# ─── Sheet 1: Course Overview ──────────────────────────────────────────────


def create_course_overview(wb, data):
    """Sheet 1: Master catalog of all available courses."""
    ws = wb.create_sheet("Course Overview")
    ws.sheet_properties.tabColor = COLOR_DARK_BLUE

    headers = [
        "Course Code", "Course Name", "Description", "Credits", "Credit System",
        "Course Type", "Department", "Faculty", "Level", "Prerequisites",
        "Corequisites", "Antirequisites", "Duration (Weeks)", "Contact Hours/Week",
        "Assessment Types", "Enrolment Cap", "Enrolment Difficulty", "Student Rating",
        "Handbook URL", "Available Semesters", "i18n Key", "Last Updated", "Is Active"
    ]
    ws.append(headers)
    style_header_row(ws)
    ws.freeze_panes = "A2"

    set_column_widths(ws, {
        "A": 15, "B": 35, "C": 50, "D": 10, "E": 20,
        "F": 18, "G": 30, "H": 25, "I": 15, "J": 25,
        "K": 25, "L": 25, "M": 15, "N": 15, "O": 25,
        "P": 12, "Q": 18, "R": 12, "S": 40, "T": 15,
        "U": 15, "V": 20, "W": 10
    })
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    courses = []
    if data and "course_overview" in data and "courses" in data.get("course_overview", {}):
        courses = data["course_overview"]["courses"]

    for c in courses:
        row = [
            safe_val(c.get("course_code")),
            safe_val(c.get("course_name")),
            safe_val(c.get("description", "")),
            c.get("credits", ""),
            safe_val(c.get("credit_system")),
            safe_val(c.get("course_type")),
            safe_val(c.get("department")),
            safe_val(c.get("faculty")),
            safe_val(c.get("level")),
            safe_val(c.get("prerequisites", ""), "None"),
            safe_val(c.get("corequisites", ""), "None"),
            safe_val(c.get("antirequisites", ""), "None"),
            c.get("duration_weeks", ""),
            c.get("contact_hours_per_week", ""),
            safe_val(c.get("assessment_types", "")),
            c.get("enrolment_cap", ""),
            safe_val(c.get("enrolment_difficulty")),
            c.get("student_rating", ""),
            safe_val(c.get("handbook_url", "")),
            safe_val(c.get("available_semesters")),
            safe_val(c.get("i18n_key", "")),
            c.get("last_updated", datetime.now().isoformat()),
            c.get("is_active", True)
        ]
        ws.append(row)

    # Conditional formatting for Course Type
    if ws.max_row > 1:
        range_f = f"F2:F{ws.max_row}"
        ws.conditional_formatting.add(range_f, CellIsRule(
            operator="equal", formula=['"Required"'],
            fill=make_fill(COLOR_DARK_RED), font=make_font(COLOR_WHITE, True)
        ))
        ws.conditional_formatting.add(range_f, CellIsRule(
            operator="equal", formula=['"Major Elective"'],
            fill=make_fill(COLOR_GREEN), font=make_font(COLOR_WHITE)
        ))

        # Enrolment Difficulty formatting
        range_q = f"Q2:Q{ws.max_row}"
        ws.conditional_formatting.add(range_q, CellIsRule(
            operator="equal", formula=['"Very Hard"'],
            border=make_border("medium", COLOR_RED),
            font=make_font(bold=True)
        ))
        ws.conditional_formatting.add(range_q, CellIsRule(
            operator="equal", formula=['"Hard"'],
            border=make_border("medium", COLOR_ORANGE)
        ))


# ─── Sheet 2: Degree Planner ──────────────────────────────────────────────


def create_degree_planner(wb, data):
    """Sheet 2: Requirement tracker + semester enrolment plan."""
    ws = wb.create_sheet("Degree Planner")
    ws.sheet_properties.tabColor = COLOR_TEAL

    # Section A: Degree Requirements
    headers_a = [
        "Requirement ID", "Requirement Category", "Requirement Name",
        "Min Credits Required", "Credits Completed", "Credits Remaining",
        "Status", "Mandatory Courses", "Notes"
    ]
    ws.append(headers_a)
    style_header_row(ws, cols=len(headers_a))
    ws.freeze_panes = "A2"

    set_column_widths(ws, {
        "A": 25, "B": 20, "C": 30, "D": 18, "E": 18,
        "F": 18, "G": 15, "H": 40, "I": 40
    })

    # Add placeholder requirement rows from recommendations if available
    recs = data.get("recommendations", {}).get("recommendations", []) if data else []
    req_categories = {}
    for r in recs:
        ct = r.get("recommendation_type", "Requirement Fulfilment")
        if ct not in req_categories:
            req_categories[ct] = []
        req_categories[ct].append(r.get("course_code", ""))

    row_idx = 2
    for cat, courses in req_categories.items():
        ws.cell(row=row_idx, column=1, value=f"REQ-{cat[:10].upper()}")
        ws.cell(row=row_idx, column=2, value=cat)
        ws.cell(row=row_idx, column=3, value=f"{cat} Courses")
        ws.cell(row=row_idx, column=4, value=0)
        ws.cell(row=row_idx, column=5, value=0)
        ws.cell(row=row_idx, column=6, value=0)
        ws.cell(row=row_idx, column=7, value="Not Started")
        ws.cell(row=row_idx, column=8, value="; ".join(courses[:10]))
        ws.cell(row=row_idx, column=9, value="")
        row_idx += 1

    if not req_categories:
        ws.append(["REQ-PLACEHOLDER", "Core", "Core Requirements", 0, 0, 0, "Not Started", "", "Add your degree requirements here"])
        row_idx += 1

    # Blank rows
    ws.append([])
    ws.append([])

    # Section B: Semester Enrolment Plan
    section_b_start = ws.max_row + 1
    headers_b = [
        "Plan ID", "Course Code", "Requirement ID", "Semester Code",
        "Status", "Priority", "Credits Earned", "Grade", "Grade Points", "Notes"
    ]
    ws.append(headers_b)
    style_header_row(ws, row=section_b_start, cols=len(headers_b))

    # Add enrolled courses from schedule
    schedule = data.get("weekly_schedule", {}).get("schedule", []) if data else []
    semester_code = data.get("meta", {}).get("semester_info", "").split(",")[0].strip() if data else ""

    seen_courses = set()
    for s in schedule:
        cc = s.get("course_code", "")
        if cc and cc not in seen_courses:
            seen_courses.add(cc)
            ws.append([
                f"PLAN-{semester_code}-{cc[:8]}",
                cc,
                "",
                semester_code,
                "Planned",
                "Medium",
                0,
                "",
                "",
                ""
            ])

    # Conditional formatting for status
    if ws.max_row > section_b_start:
        status_range = f"E{section_b_start+1}:E{ws.max_row}"
        ws.conditional_formatting.add(status_range, CellIsRule(
            operator="equal", formula=['"Completed"'],
            fill=make_fill(COLOR_GREEN_FILL), font=make_font(COLOR_WHITE)
        ))
        ws.conditional_formatting.add(status_range, CellIsRule(
            operator="equal", formula=['"Planned"'],
            fill=make_fill(COLOR_BLUE), font=make_font(COLOR_WHITE)
        ))
        ws.conditional_formatting.add(status_range, CellIsRule(
            operator="equal", formula=['"Failed"'],
            fill=make_fill(COLOR_RED), font=make_font(COLOR_WHITE, True)
        ))


# ─── Sheet 3: AI Recommendations ───────────────────────────────────────────


def create_ai_recommendations(wb, data):
    """Sheet 3: AI-generated course suggestions."""
    ws = wb.create_sheet("AI Recommendations")
    ws.sheet_properties.tabColor = COLOR_CYAN

    headers = [
        "Recommendation ID", "Course Code", "Reason", "Recommendation Type",
        "Confidence Score", "Priority", "Recommended Semester", "Alternatives",
        "Decision", "Decision Timestamp", "Resulting Plan ID", "Generated At",
        "Expires At"
    ]
    ws.append(headers)
    style_header_row(ws)
    ws.freeze_panes = "A2"

    set_column_widths(ws, {
        "A": 25, "B": 15, "C": 60, "D": 25, "E": 15,
        "F": 12, "G": 15, "H": 30, "I": 12, "J": 20,
        "K": 25, "L": 20, "M": 20
    })
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    recs = []
    if data and "recommendations" in data:
        recs = data["recommendations"].get("recommendations", [])

    now = datetime.now().isoformat()
    for i, r in enumerate(recs, 1):
        ws.append([
            f"REC-{now[:10].replace('-', '')}-{i:04d}",
            safe_val(r.get("course_code")),
            safe_val(r.get("reason", "")),
            safe_val(r.get("recommendation_type")),
            r.get("confidence_score", 0),
            safe_val(r.get("priority")),
            safe_val(r.get("recommended_semester", "")),
            safe_val(r.get("alternatives", ""), "None"),
            "Pending",
            "",
            "",
            now,
            ""
        ])

    # Conditional formatting
    if ws.max_row > 1:
        conf_range = f"E2:E{ws.max_row}"
        ws.conditional_formatting.add(conf_range, CellIsRule(
            operator="greaterThanOrEqual", formula=["0.9"],
            fill=make_fill(COLOR_GOLD)
        ))
        ws.conditional_formatting.add(conf_range, CellIsRule(
            operator="lessThan", formula=["0.5"],
            fill=make_fill(COLOR_LIGHT_RED)
        ))

        prio_range = f"F2:F{ws.max_row}"
        ws.conditional_formatting.add(prio_range, CellIsRule(
            operator="equal", formula=['"Critical"'],
            fill=make_fill(COLOR_DARK_RED), font=make_font(COLOR_WHITE, True)
        ))


# ─── Sheet 4: Weekly Timetable ─────────────────────────────────────────────


def create_weekly_timetable(wb, data):
    """Sheet 4: Visual weekly grid."""
    ws = wb.create_sheet("Weekly Timetable")
    ws.sheet_properties.tabColor = COLOR_PURPLE

    days = ["Time", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    ws.append(days)
    style_header_row(ws, cols=len(days))
    ws.freeze_panes = "B2"

    # Time slots: 08:00 to 21:30 in 30-min increments
    time_slots = []
    for h in range(8, 22):
        time_slots.append(f"{h:02d}:00")
        time_slots.append(f"{h:02d}:30")
    time_slots.append("22:00")

    for ts in time_slots:
        ws.append([ts] + [""] * 7)

    set_column_widths(ws, {
        "A": 10, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22, "H": 22
    })

    # Populate schedule
    schedule = []
    if data and "weekly_schedule" in data:
        schedule = data["weekly_schedule"].get("schedule", [])

    # Build a lookup: day -> list of (start, end, course_code, session_type, location, instructor)
    day_sessions = {}
    for s in schedule:
        day = s.get("day", "")
        if day not in day_sessions:
            day_sessions[day] = []
        day_sessions[day].append(s)

    # Map days to columns
    day_cols = {
        "Monday": 2, "Tuesday": 3, "Wednesday": 4, "Thursday": 5,
        "Friday": 6, "Saturday": 7, "Sunday": 8
    }

    course_types = {}
    if data and "course_overview" in data:
        for c in data["course_overview"].get("courses", []):
            course_types[c.get("course_code", "")] = c.get("course_type", "")

    for day, sessions in day_sessions.items():
        col = day_cols.get(day)
        if col is None:
            continue
        for s in sessions:
            start = s.get("start_time", "")
            end = s.get("end_time", "")
            cc = s.get("course_code", "")
            st = s.get("session_type", "")
            campus = s.get("campus", "")
            building = s.get("building", "")
            room = s.get("room", "")
            instructor = s.get("instructor", "")

            location = " / ".join(filter(lambda x: x and x != NOT_FOUND, [campus, building, room]))
            if not location:
                location = NOT_FOUND

            cell_text = f"{cc}\n{st}\n{location}"
            if instructor and instructor != NOT_FOUND:
                cell_text += f"\n{instructor}"

            # Find the time slot rows
            start_row = None
            end_row = None
            for i, ts in enumerate(time_slots, 2):
                if ts == start:
                    start_row = i
                if ts == end:
                    end_row = i
                    break

            if start_row and end_row:
                # Write to the first cell
                cell = ws.cell(row=start_row, column=col, value=cell_text)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = make_border("thin", "808080")

                # Apply coloring based on course type
                ct = course_types.get(cc, "")
                if ct == "Required":
                    cell.fill = make_fill(COLOR_DARK_RED)
                    cell.font = make_font(COLOR_WHITE, True)
                elif ct == "Major Elective":
                    cell.fill = make_fill(COLOR_GREEN)
                    cell.font = make_font(COLOR_WHITE)
                elif ct == "General Education":
                    cell.fill = make_fill(COLOR_BLUE)
                    cell.font = make_font(COLOR_WHITE)

                # Session type coloring
                if "Lab" in st:
                    cell.fill = make_fill(COLOR_PURPLE)
                    cell.font = make_font(COLOR_WHITE)
                elif "Tutorial" in st:
                    cell.fill = make_fill(COLOR_LIGHT_BLUE)
                elif "Workshop" in st:
                    cell.fill = make_fill(COLOR_ORANGE)
                    cell.font = make_font(COLOR_WHITE)

                # Merge cells if spanning multiple time slots
                if end_row > start_row + 1:
                    ws.merge_cells(
                        start_row=start_row, start_column=col,
                        end_row=end_row - 1, end_column=col
                    )

    # Free time grey fill for empty cells
    for row in range(2, ws.max_row + 1):
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)
            if cell.value is None or cell.value == "":
                cell.fill = make_fill(COLOR_GREY)

    # Row heights
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 25


# ─── Sheet 5: Academic Calendar ────────────────────────────────────────────


def create_academic_calendar(wb, data):
    """Sheet 5: Institution-wide academic dates."""
    ws = wb.create_sheet("Academic Calendar")
    ws.sheet_properties.tabColor = COLOR_ORANGE

    headers = [
        "Event ID", "Event Name", "Event Type", "Start Date", "End Date",
        "Start Time", "End Time", "Semester Code", "Is Recurring",
        "Recurrence Rule", "Location", "Description", "Calendar UID",
        "Reminder (Minutes)"
    ]
    ws.append(headers)
    style_header_row(ws)
    ws.freeze_panes = "A2"

    set_column_widths(ws, {
        "A": 25, "B": 35, "C": 20, "D": 15, "E": 15,
        "F": 10, "G": 10, "H": 15, "I": 12, "J": 40,
        "K": 20, "L": 50, "M": 35, "N": 15
    })
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Try to extract academic calendar from semester_info
    meta = data.get("meta", {}) if data else {}
    semester_info = meta.get("semester_info", "")
    semester_code = semester_info.split(",")[0].strip() if semester_info else ""

    # Add basic semester events
    if semester_info:
        parts = [p.strip() for p in semester_info.split(",")]
        ws.append([
            f"EVT-{semester_code}-START",
            f"{semester_code} Semester Start",
            "Semester Start",
            parts[1].replace("starts ", "").strip() if len(parts) > 1 else "",
            parts[1].replace("starts ", "").strip() if len(parts) > 1 else "",
            "", "", semester_code, False, "", "", "Semester start date", "", 0
        ])
        if len(parts) > 2:
            ws.append([
                f"EVT-{semester_code}-END",
                f"{semester_code} Semester End",
                "Semester End",
                parts[2].replace("ends ", "").strip(),
                parts[2].replace("ends ", "").strip(),
                "", "", semester_code, False, "", "", "Semester end date", "", 0
            ])

    # Conditional formatting
    if ws.max_row > 1:
        etype_range = f"C2:C{ws.max_row}"
        ws.conditional_formatting.add(etype_range, CellIsRule(
            operator="equal", formula=['"Holiday"'],
            fill=make_fill(COLOR_YELLOW)
        ))
        ws.conditional_formatting.add(etype_range, CellIsRule(
            operator="equal", formula=['"Exam Period"'],
            fill=make_fill(COLOR_RED), font=make_font(COLOR_WHITE)
        ))
        ws.conditional_formatting.add(etype_range, CellIsRule(
            operator="equal", formula=['"Break"'],
            fill=make_fill("D6EAF8")
        ))


# ─── Sheet 6: Raw Schedule Database ─────────────────────────────────────────


def create_raw_schedule(wb, data):
    """Sheet 6: Normalized session database (hidden by default)."""
    ws = wb.create_sheet("Raw Schedule Database")
    ws.sheet_properties.tabColor = COLOR_DARK_GREY
    ws.sheet_state = "hidden"

    headers = [
        "Session ID", "Course Code", "Session Type", "Day", "Start Time",
        "End Time", "Campus", "Building", "Room", "Instructor",
        "Week Pattern", "Start Week", "End Week", "Effective From",
        "Effective Until", "Notes"
    ]
    ws.append(headers)
    style_header_row(ws)
    ws.freeze_panes = "A2"

    set_column_widths(ws, {
        "A": 25, "B": 15, "C": 15, "D": 12, "E": 10,
        "F": 10, "G": 15, "H": 20, "I": 10, "J": 20,
        "K": 15, "L": 10, "M": 10, "N": 15, "O": 15, "P": 40
    })
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    schedule = []
    if data and "weekly_schedule" in data:
        schedule = data["weekly_schedule"].get("schedule", [])

    for s in schedule:
        ws.append([
            safe_val(s.get("session_id")),
            safe_val(s.get("course_code")),
            safe_val(s.get("session_type")),
            safe_val(s.get("day")),
            safe_val(s.get("start_time")),
            safe_val(s.get("end_time")),
            safe_val(s.get("campus", "")),
            safe_val(s.get("building", "")),
            safe_val(s.get("room", "")),
            safe_val(s.get("instructor", "")),
            safe_val(s.get("week_pattern")),
            s.get("start_week", ""),
            s.get("end_week", ""),
            safe_val(s.get("effective_from", "")),
            safe_val(s.get("effective_until", "")),
            ""
        ])


# ─── Sheet 7: Degree Audit & Risk Assessment (cloud-only) ─────────────────────

def create_degree_audit(wb, data):
    """Create the Degree Audit & Risk Assessment sheet (cloud-only, additive).
    Populated from combined.degree_audit (cloud stp_deg_audit output).
    If absent, creates headers only — never breaks existing 6-sheet behavior."""
    ws = wb.create_sheet("Degree Audit & Risk Assessment")
    audit = data.get("degree_audit", {}) or {}

    headers = ["Item", "Detail", "Severity"]
    ws.append(headers)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8B0000")

    # Credit progress
    total = audit.get("total_credits_planned")
    note = audit.get("credit_progress_note", "—")
    ws.append(["Credit Progress", f"Total planned: {total} | {note}" if total else note, ""])

    # Prerequisite issues
    ws.append(["Prerequisite Issues", "", ""])
    for issue in audit.get("prerequisite_issues", []):
        ws.append([issue.get("course_code", ""),
                   "Missing prereqs: " + ", ".join(issue.get("missing_prereqs", [])),
                   issue.get("severity", "Medium")])

    # Overload weeks
    ws.append(["Overload Weeks", "", ""])
    for ow in audit.get("overload_weeks", []):
        ws.append(["Week " + str(ow.get("week", "")),
                   ", ".join(ow.get("courses", [])) + (" | " + ow.get("risk", "") if ow.get("risk") else ""),
                   "High"])

    # Sequence alerts
    ws.append(["Sequence Alerts", "", ""])
    for sa in audit.get("sequence_alerts", []):
        ws.append([sa.get("course_code", ""), sa.get("alert", ""), sa.get("severity", "Medium")])

    # Summary
    if audit.get("summary"):
        ws.append(["Summary", audit.get("summary"), ""])

    # Color severity
    for row in ws.iter_rows(min_row=2):
        sev = row[2].value if len(row) > 2 else None
        if sev == "High":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FCEBEB")
        elif sev == "Medium":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FAEEDA")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12


# ─── Main ──────────────────────────────────────────────────────────────────


def generate_workbook(input_path, output_path):
    """Generate the complete Excel workbook from JSON input."""
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets in order
    create_course_overview(wb, data)
    create_degree_planner(wb, data)
    create_ai_recommendations(wb, data)
    create_weekly_timetable(wb, data)
    create_academic_calendar(wb, data)
    create_raw_schedule(wb, data)
    # Sheet 7: Degree Audit (cloud-only, additive; headers-only if no data)
    create_degree_audit(wb, data)

    # Save
    wb.save(output_path)
    print(f"Excel workbook generated: {output_path}")
    print(f"Sheets: Course Overview, Degree Planner, AI Recommendations, "
          f"Weekly Timetable, Academic Calendar, Raw Schedule Database, Degree Audit & Risk Assessment")

    # Summary
    total_courses = len(data.get("course_overview", {}).get("courses", []))
    total_sessions = len(data.get("weekly_schedule", {}).get("schedule", []))
    total_recs = len(data.get("recommendations", {}).get("recommendations", []))
    print(f"Courses: {total_courses} | Sessions: {total_sessions} | Recommendations: {total_recs}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel workbook from course planner JSON data."
    )
    parser.add_argument("input", help="Path to the input JSON file")
    parser.add_argument("--output", "-o", default="course_schedule.xlsx",
                        help="Output Excel file path (default: course_schedule.xlsx)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)

    generate_workbook(args.input, args.output)


if __name__ == "__main__":
    main()
