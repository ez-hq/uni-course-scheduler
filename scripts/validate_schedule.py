#!/usr/bin/env python3
"""
University Course Planner — Schedule Validator

Validates a generated Excel workbook against the local validation contract.
Checks: time conflicts, lunch break preservation, daily density, credit totals,
required course coverage, and missing field detection.

Usage:
    python3 validate_schedule.py <excel_file.xlsx>
"""

import argparse
import sys
from datetime import datetime, timedelta

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

NOT_FOUND = "— (not found, please add)"


def parse_time(time_str):
    """Parse HH:MM string into minutes since midnight."""
    if not time_str or time_str == NOT_FOUND:
        return None
    try:
        parts = str(time_str).strip().split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except (ValueError, IndexError):
        return None


def check_time_conflicts(schedule):
    """Check no two sessions overlap on the same day."""
    issues = []
    by_day = {}
    for s in schedule:
        day = s.get("day", "")
        if not day:
            continue
        if day not in by_day:
            by_day[day] = []
        by_day[day].append(s)

    for day, sessions in by_day.items():
        sessions.sort(key=lambda x: parse_time(x.get("start_time", "")) or 0)
        for i in range(len(sessions)):
            for j in range(i + 1, len(sessions)):
                s1_start = parse_time(sessions[i].get("start_time", ""))
                s1_end = parse_time(sessions[i].get("end_time", ""))
                s2_start = parse_time(sessions[j].get("start_time", ""))
                s2_end = parse_time(sessions[j].get("end_time", ""))
                if None in (s1_start, s1_end, s2_start, s2_end):
                    continue
                if s1_start < s2_end and s2_start < s1_end:
                    issues.append(
                        f"{day}: {sessions[i].get('course_code', '')} "
                        f"({sessions[i].get('start_time', '')}-{sessions[i].get('end_time', '')}) "
                        f"conflicts with {sessions[j].get('course_code', '')} "
                        f"({sessions[j].get('start_time', '')}-{sessions[j].get('end_time', '')})"
                    )
    return len(issues) == 0, issues


def check_lunch_break(schedule):
    """Check 12:00-13:00 is free every day."""
    issues = []
    by_day = {}
    for s in schedule:
        day = s.get("day", "")
        if not day:
            continue
        if day not in by_day:
            by_day[day] = []
        by_day[day].append(s)

    for day, sessions in by_day.items():
        for s in sessions:
            start = parse_time(s.get("start_time", ""))
            end = parse_time(s.get("end_time", ""))
            if start is None or end is None:
                continue
            if start < 780 and end > 720:
                issues.append(f"{day}: {s.get('course_code', '')} overlaps lunch break (12:00-13:00)")
    return len(issues) == 0, issues


def check_daily_density(schedule):
    """Check no day exceeds 6 hours of class."""
    issues = []
    by_day = {}
    for s in schedule:
        day = s.get("day", "")
        if not day:
            continue
        if day not in by_day:
            by_day[day] = []
        by_day[day].append(s)

    for day, sessions in by_day.items():
        total_minutes = 0
        for s in sessions:
            start = parse_time(s.get("start_time", ""))
            end = parse_time(s.get("end_time", ""))
            if start is not None and end is not None:
                total_minutes += end - start
        total_hours = total_minutes / 60
        if total_hours > 6:
            issues.append(f"{day}: {total_hours:.1f} hours exceeds 6-hour limit")
    return len(issues) == 0, issues


def check_credit_total(courses):
    """Check total credits are reasonable."""
    total = 0
    for c in courses:
        credits = c.get("credits", 0)
        try:
            total += float(credits)
        except (ValueError, TypeError):
            continue
    if total == 0:
        return True, ["No credits found (acceptable for empty catalog)"]
    if total < 10 or total > 80:
        return False, [f"Total credits {total} outside typical range (10-80)"]
    return True, [f"Total credits: {total}"]


def check_required_coverage(courses):
    """Check at least one Required course exists."""
    required = [c for c in courses if c.get("course_type") == "Required"]
    if not required:
        return False, ["No Required courses found in catalog"]
    return True, [f"{len(required)} required course(s) in catalog"]


def check_missing_fields(courses, schedule):
    """Check no more than 10% of critical fields are NOT_FOUND."""
    critical_fields = [
        "course_code", "course_name", "credits", "course_type",
        "day", "start_time", "end_time"
    ]
    total_fields = 0
    missing_fields = 0
    for c in courses:
        for field in critical_fields[:4]:
            total_fields += 1
            val = c.get(field, "")
            if not val or val == "NOT_FOUND" or val == NOT_FOUND:
                missing_fields += 1
    for s in schedule:
        for field in critical_fields[4:]:
            total_fields += 1
            val = s.get(field, "")
            if not val or val == "NOT_FOUND" or val == NOT_FOUND:
                missing_fields += 1

    if total_fields == 0:
        return True, ["No fields to check"]
    pct = (missing_fields / total_fields) * 100
    if pct > 10:
        return False, [f"{missing_fields}/{total_fields} fields missing ({pct:.1f}%)"]
    return True, [f"{missing_fields}/{total_fields} fields missing ({pct:.1f}%)"]


def main():
    parser = argparse.ArgumentParser(description="Validate course schedule Excel workbook.")
    parser.add_argument("excel_file", help="Path to the .xlsx file to validate")
    args = parser.parse_args()

    wb = load_workbook(args.excel_file, data_only=True)

    schedule = []
    if "Raw Schedule Database" in wb.sheetnames:
        ws = wb["Raw Schedule Database"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            entry = {}
            for i, h in enumerate(headers):
                if h and i < len(row):
                    entry[h.lower().replace(" ", "_").replace("/", "_")] = row[i]
            schedule.append(entry)

    courses = []
    if "Course Overview" in wb.sheetnames:
        ws = wb["Course Overview"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            entry = {}
            for i, h in enumerate(headers):
                if h and i < len(row):
                    entry[h.lower().replace(" ", "_").replace("/", "_")] = row[i]
            courses.append(entry)

    checks = [
        ("time_conflict", "Time Conflict Detection", *check_time_conflicts(schedule)),
        ("lunch_break", "Lunch Break Preservation (12:00-13:00)", *check_lunch_break(schedule)),
        ("daily_density", "Daily Course Density (max 6h)", *check_daily_density(schedule)),
        ("credit_total", "Credit Total Validation", *check_credit_total(courses)),
        ("required_coverage", "Required Course Coverage", *check_required_coverage(courses)),
        ("missing_fields", "Missing Field Detection", *check_missing_fields(courses, schedule)),
    ]

    all_pass = all(result for _, _, result, _ in checks)
    passed = sum(1 for _, _, result, _ in checks if result)

    report = []
    report.append("# Schedule Validation Report")
    report.append("")
    report.append(f"**Generated:** {datetime.now().isoformat()}")
    report.append(f"**Overall:** {'PASS' if all_pass else 'FAIL'}")
    report.append(f"**Checks:** {passed}/{len(checks)} passed")
    report.append("")
    report.append("## Checks")
    report.append("")
    report.append("| ID | Name | Result | Details |")
    report.append("|----|------|--------|---------|")
    for check_id, name, result, details in checks:
        status = "PASS" if result else "FAIL"
        detail = "; ".join(details) if details else ""
        report.append(f"| {check_id} | {name} | {status} | {detail} |")

    print("\n".join(report))
    return 0 if all_pass else 1


if __name__ == "__main__":
    raise SystemExit(main())
